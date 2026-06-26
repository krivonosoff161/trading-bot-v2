"""
build_journal.py — trading journal Excel.

Sheets: Скринер | Симулятор | Main WS | BB Fade | Памп | Импульс | Ручные | Дашборд | Графики

Импульс — paper impulse pump (рывок). Метрика — net_pct (ride-the-move), не TP/SL.

Run via update_journal.bat after labelers.
"""
import asyncio
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from dotenv import load_dotenv
load_dotenv()

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
except ImportError:
    print("openpyxl не установлен: pip install openpyxl")
    sys.exit(1)

_ROOT               = Path(__file__).resolve().parent.parent
SIGNAL_LOG          = _ROOT / "logs" / "signals" / "signal_log.jsonl"
SIGNAL_LABELS       = _ROOT / "logs" / "signals" / "signal_labels.jsonl"
PUMP_SIGNALS_LOG    = _ROOT / "logs" / "pump" / "pump_signals.jsonl"
PUMP_LABELS_LOG     = _ROOT / "logs" / "pump" / "pump_labels.jsonl"
MAIN_SIGNALS_LOG    = _ROOT / "logs" / "signals" / "main_signals.jsonl"
MAIN_SIGNALS_LABELS = _ROOT / "logs" / "signals" / "main_signals_labels.jsonl"
BB_FADE_SIGNALS_LOG = _ROOT / "logs" / "bb_fade" / "bb_fade_signals.jsonl"
IMPULSE_SIGNALS_LOG = _ROOT / "logs" / "impulse_pump" / "impulse_pump_signals.jsonl"
IMPULSE_OUTCOMES_LOG = _ROOT / "logs" / "impulse_pump" / "impulse_pump_outcomes.jsonl"
MAIN_IMPULSE_SIGNALS_LOG = _ROOT / "logs" / "main_impulse" / "main_impulse_signals.jsonl"
MAIN_IMPULSE_OUTCOMES_LOG = _ROOT / "logs" / "main_impulse" / "main_impulse_outcomes.jsonl"
JOURNAL_PATH        = Path(__file__).parent / "journal.xlsx"

# Reset boundary (27.05.2026): все pre-reset торговые данные перенесены сюда = СТАРАЯ стратегия (архив).
# Новый запуск пишет свежее в logs/. Журнал мерджит оба источника (дедуп по signal_id).
RESET_DATE          = "27.05.2026"
RESET_ARCHIVE       = _ROOT / "logs_archive" / "pre-reset_2026-05-27"

# Archive paths — same schema as current logs, just historical
ARCHIVE_DIRS        = [
    _ROOT / "logs_archive" / "09.05.2026" / "signals",
    _ROOT / "logs_archive" / "signals",
    _ROOT / "logs_archive" / "09.05.2026" / "pump",
    RESET_ARCHIVE / "signals",
]

# Data roots для single-file лоадеров: живые logs/ (новое) + reset-архив (история).
_DATA_ROOTS = [_ROOT / "logs", RESET_ARCHIVE]


def _private_research_root() -> Path:
    """Private Strategy Lab root for derived paper artifacts."""
    return Path(os.getenv(
        "TRADING_BOT_RESEARCH_ROOT",
        str(Path.home() / "github_projects" / "trading-bot-research" / "strategy-lab"),
    )).expanduser()


def _paper_signal_training_path() -> Path:
    return _private_research_root() / "state" / "derived" / "paper_signal_training.jsonl"


def _jsonl_multi(rel: str) -> list:
    """Читает относительный jsonl из всех data-roots (live logs/ + reset-архив), конкатенация.

    Дедуп — на стороне лоадеров (по signal_id). _load_jsonl резолвится в рантайме.
    """
    out: list = []
    for root in _DATA_ROOTS:
        out += _load_jsonl(root / rel)
    return out

# ── Palette ──────────────────────────────────────────────────────────────────
def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)

F_HDR    = _fill("1F4E79")   # dark blue   — data sheet headers
F_INPUT  = _fill("FFF2CC")   # yellow      — user-editable cells
F_TP     = _fill("C6EFCE")   # green
F_SL     = _fill("FFC7CE")   # red
F_TIME   = _fill("FFEB9C")   # amber
F_SUM    = _fill("D9E1F2")   # light blue  — summary rows
F_DASH_T = _fill("2E4057")   # dark navy   — dashboard section titles
F_DASH_A = _fill("4472C4")   # blue        — dashboard column headers
F_STRIPE = _fill("F2F2F2")   # light grey  — alternating rows

FONT_HDR   = Font(color="FFFFFF", bold=True, size=10)
FONT_BOLD  = Font(bold=True)
FONT_DTIT  = Font(color="FFFFFF", bold=True, size=11)
FONT_DHDR  = Font(color="FFFFFF", bold=True, size=10)
ALIGN_C    = Alignment(horizontal="center")

OUTCOME_FILL = {
    "TP": F_TP, "TP1": F_TP, "TP2": F_TP,
    "SL": F_SL, "STOP": F_SL,
    "TIME": F_TIME, "TIME_EXIT": F_TIME,
}


# ── Generic helpers ───────────────────────────────────────────────────────────

def _load_jsonl(path: Path) -> list:
    if not path.exists():
        return []
    out = []
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line:
            try:
                out.append(json.loads(line))
            except json.JSONDecodeError:
                pass
    return out


def _load_paper_watch() -> list:
    """Load derived paper-watch outcomes from the private research root.

    This is intentionally a read-only JSONL loader. It does not use exchange clients,
    account endpoints, Telegram, or LLM providers.
    """
    rows = []
    for row in _load_jsonl(_paper_signal_training_path()):
        if row.get("schema") != "PaperSignalTrainingRow.v1":
            continue
        rows.append({
            "created_at": row.get("created_at", ""),
            "symbol": row.get("symbol", ""),
            "okx_inst_id": row.get("okx_inst_id", ""),
            "timeframe": row.get("timeframe", ""),
            "family": row.get("family", ""),
            "side": row.get("side", ""),
            "exit_mode": row.get("exit_mode", ""),
            "status": row.get("status", ""),
            "result": row.get("result", ""),
            "net_pct": row.get("net_pct"),
            "net_r": row.get("net_r"),
            "mfe_pct": row.get("mfe_pct"),
            "mae_pct": row.get("mae_pct"),
            "capture": row.get("capture"),
            "diagnosis": row.get("diagnosis", ""),
            "risk_pct": row.get("risk_pct"),
            "max_hold_minutes": row.get("max_hold_minutes"),
            "source": row.get("source", ""),
            "paper_only": bool(row.get("paper_only")),
        })
    rows.sort(key=lambda x: str(x.get("created_at") or ""), reverse=True)
    return rows


def _fmt_dt(ts_ms) -> str:
    if not ts_ms:
        return ""
    return datetime.fromtimestamp(int(ts_ms) / 1000, tz=timezone.utc).strftime("%d.%m %H:%M")


def _set_widths(ws, widths: list) -> None:
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _hdr(ws, row: int, cols: list, fill=None, font=None) -> None:
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill  = fill or F_HDR
        cell.font  = font or FONT_HDR
        cell.alignment = ALIGN_C


def _dash_title(ws, row: int, col: int, text: str, span: int) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = F_DASH_T
    cell.font = FONT_DTIT
    cell.alignment = ALIGN_C
    if span > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=col + span - 1,
        )


def _dash_hdr(ws, row: int, col_start: int, cols: list) -> None:
    for i, h in enumerate(cols):
        cell = ws.cell(row=row, column=col_start + i, value=h)
        cell.fill = F_DASH_A
        cell.font = FONT_DHDR
        cell.alignment = ALIGN_C


def _dw(ws, row: int, col: int, value, fill=None, font=None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font


# ── Data loaders ─────────────────────────────────────────────────────────────

def _load_screener() -> list:
    """Загружает signal_log + signal_labels из текущих и архивных папок.

    Схема одинаковая — мерджим по signal_id, дедупликация автоматическая.
    Архивные файлы (logs_archive/) дополняют живые данные историческими.
    """
    # Все source файлы (схема идентична)
    signal_files = [SIGNAL_LOG]
    label_files = [SIGNAL_LABELS]
    for d in ARCHIVE_DIRS:
        if not d.exists():
            continue
        for p in d.glob("signal_log*.jsonl"):
            signal_files.append(p)
        for p in d.glob("signal_labels*.jsonl"):
            label_files.append(p)

    # Мердж labels по signal_id
    labels: dict = {}
    for f in label_files:
        for r in _load_jsonl(f):
            sid = r.get("signal_id", "")
            if sid and sid not in labels:
                labels[sid] = r

    # Мердж signals по signal_id (если дубликат — оставляем первый)
    seen_ids: set = set()
    raw_signals: list = []
    for f in signal_files:
        for r in _load_jsonl(f):
            sid = r.get("signal_id", "")
            if not sid or sid in seen_ids:
                continue
            seen_ids.add(sid)
            raw_signals.append(r)

    out = []
    for sig in raw_signals:
        lab   = labels.get(sig.get("signal_id", ""), {})
        style = sig.get("style") or ("FADE" if sig.get("source") == "bb_fade" else "FAST")
        out.append({
            "dt":       _fmt_dt(sig.get("ts_ms")),
            "pair":     sig.get("symbol", ""),
            "style":    style,
            "regime":   sig.get("regime", ""),
            "side":     "LONG" if sig.get("side") == "buy" else "SHORT",
            "entry":    sig.get("close"),
            "sl":       sig.get("sl"),
            "tp":       sig.get("tp"),
            "hold_min": sig.get("max_hold_min"),
            "outcome":  lab.get("outcome", ""),
            "R":        lab.get("exit_r"),
            "MFE_R":    lab.get("mfe_r"),
            "MAE_R":    lab.get("mae_r"),
            "adx_1h":   sig.get("adx_1h"),
            "slope_1h": sig.get("slope_1h"),
            "slope_15m":sig.get("slope_15m"),
            "funding":  sig.get("funding") or 0.0,
            "ts_ms":    sig.get("ts_ms", 0),  # для сортировки
        })
    # Сортируем по времени, новейшие сверху
    out.sort(key=lambda x: x.get("ts_ms", 0), reverse=True)
    return out


def _calc_r_to_tp(entry, sl, tp, side: str):
    """R-multiple of TP relative to SL distance. Returns None if invalid."""
    if entry is None or sl is None or tp is None:
        return None
    try:
        entry = float(entry)
        sl = float(sl)
        tp = float(tp)
    except (TypeError, ValueError):
        return None
    sl_dist = abs(entry - sl)
    if sl_dist == 0:
        return None
    if side == "LONG" or side == "buy":
        return round((tp - entry) / sl_dist, 2) if tp > entry else None
    return round((entry - tp) / sl_dist, 2) if tp < entry else None


# ── Main WS execution-cost model (29.05.2026) ────────────────────────────────
# Client fills (confirmed by trader): ENTRY market/taker, EXIT TP/SL limit/maker;
# manual exits negligible. OKX SWAP ~0.05% taker / ~0.02% maker per side; market-
# entry adverse slippage ~0.03% (config.yaml). Round-trip price-cost ~0.10%.
# In R-units: cost_R = COST_PCT_RT / SL-distance%  (per signal).
FEE_TAKER_PCT = 0.05    # entry, market
FEE_MAKER_PCT = 0.02    # exit, limit TP/SL
ENTRY_SLIP_PCT = 0.03   # market-entry adverse slippage
COST_PCT_RT = FEE_TAKER_PCT + FEE_MAKER_PCT + ENTRY_SLIP_PCT  # 0.10% round-trip


def _cost_r(entry, sl) -> float:
    """Round-trip execution cost expressed in R (fraction of SL distance)."""
    try:
        entry = float(entry)
        sl = float(sl)
    except (TypeError, ValueError):
        return 0.0
    if entry == 0:
        return 0.0
    sl_dist_pct = abs(entry - sl) / abs(entry) * 100.0
    if sl_dist_pct <= 0:
        return 0.0
    return COST_PCT_RT / sl_dist_pct


def _r_from_exit(entry, sl, exit_price, side: str):
    """Signed R of an arbitrary exit price (used for TIME exits)."""
    if entry is None or sl is None or exit_price is None:
        return None
    try:
        entry = float(entry)
        sl = float(sl)
        exit_price = float(exit_price)
    except (TypeError, ValueError):
        return None
    sl_dist = abs(entry - sl)
    if sl_dist == 0:
        return None
    s = 1 if side in ("LONG", "buy") else -1
    return round(s * (exit_price - entry) / sl_dist, 2)


def _outcome_to_tp_unified(outcome: str) -> str:
    """Unify TP1/TP2 → TP (client closes at TP1 anyway)."""
    if outcome in ("TP1", "TP2"):
        return "TP"
    return outcome or ""


def _exit_r_gross(entry, sl, tp1, side: str, outcome: str, exit_price=None):
    """GROSS R (no cost): TP → +R_to_TP1 (single-TP), SL → -1, TIME → signed R at exit."""
    out_norm = _outcome_to_tp_unified(outcome)
    if out_norm == "TP":
        return _calc_r_to_tp(entry, sl, tp1, side)
    if out_norm == "SL":
        return -1.0
    if out_norm == "TIME":
        return _r_from_exit(entry, sl, exit_price, side)
    return None  # INVALID / empty


def _exit_r_at_tp1(entry, sl, tp1, side: str, outcome: str):
    """Back-compat for other channels: GROSS R, TIME->None (no cost, no exit price)."""
    return _exit_r_gross(entry, sl, tp1, side, outcome)


def _load_main_ws() -> list:
    signals = {
        (s.get("id") or s.get("signal_id", "")): s
        for s in _jsonl_multi("signals/main_signals.jsonl")
    }
    labels = {r["signal_id"]: r for r in _jsonl_multi("signals/main_signals_labels.jsonl") if r.get("signal_id")}
    out = []
    for sig_id, sig in signals.items():
        lab = labels.get(sig_id, {})
        ts  = sig.get("ts", "")
        side = "LONG" if sig.get("side") == "buy" else "SHORT"
        outcome_raw = lab.get("outcome", "")
        outcome = _outcome_to_tp_unified(outcome_raw)
        tp2_reached = outcome_raw == "TP2"  # info only — client doesn't trade TP2
        gross = _exit_r_gross(sig.get("entry"), sig.get("sl"), sig.get("tp1"), side, outcome_raw, lab.get("exit_price"))
        cost = _cost_r(sig.get("entry"), sig.get("sl"))
        exit_r = round(gross - cost, 3) if gross is not None else None  # NET R
        out.append({
            "dt":         ts[5:16].replace("T", " ") if ts else "",
            "pair":       sig.get("symbol", "").replace("-USDT-SWAP", ""),
            "regime":     sig.get("regime", ""),
            "style":      sig.get("trade_style", ""),
            "side":       side,
            "entry":      sig.get("entry"),
            "sl":         sig.get("sl"),
            "tp":         sig.get("tp1"),
            "hold_min":   sig.get("hold_min"),
            "outcome":    outcome,
            "tp2_reached": tp2_reached,
            "exit_price": lab.get("exit_price"),
            "exit_r":     exit_r,
            "exit_r_gross": gross,
        })
    out.sort(key=lambda x: x["dt"], reverse=True)
    return out


def _load_pump() -> list:
    """Загружает pump entries + labels из текущих и архивных папок."""
    signal_files = [PUMP_SIGNALS_LOG]
    label_files  = [PUMP_LABELS_LOG]
    # Archive
    for d in [_ROOT / "logs_archive" / "09.05.2026" / "pump", RESET_ARCHIVE / "pump"]:
        if not d.exists():
            continue
        for p in d.glob("pump_signals*.jsonl"):
            signal_files.append(p)
        for p in d.glob("pump_labels*.jsonl"):
            label_files.append(p)
    for d in [_ROOT / "logs_archive" / "pump"]:
        if not d.exists():
            continue
        for p in d.glob("pump_signals*.jsonl"):
            signal_files.append(p)

    # Мердж entries по signal_id
    entries: dict = {}
    for f in signal_files:
        for r in _load_jsonl(f):
            if r.get("type") != "ENTRY":
                continue
            sid = r.get("signal_id", "")
            if sid and sid not in entries:
                entries[sid] = r

    out = []
    seen_exit_ids: set = set()
    for f in label_files:
        for rec in _load_jsonl(f):
            if rec.get("type") != "EXIT":
                continue
            sid = rec.get("signal_id", "")
            if sid in seen_exit_ids:
                continue
            seen_exit_ids.add(sid)
            entry  = entries.get(sid, {})
            opened = rec.get("opened_at", "")
            sl_raw = entry.get("paper_sl")
            tp_raw = entry.get("paper_tp")
            out.append({
                "dt":      opened[:16].replace("T", " ") if opened else "",
                "pair":    rec.get("sym", "").replace("-USDT-SWAP", "").replace("-SWAP", ""),
                "entry":   rec.get("entry_price"),
                "sl":      float(sl_raw) if sl_raw else None,
                "tp":      float(tp_raw) if tp_raw else None,
                "outcome": rec.get("exit_reason", ""),
                "hold":    rec.get("hold_min"),
                "hour":    entry.get("hour_utc"),
                "net_pnl": rec.get("net_pnl_pct"),
            })
    out.sort(key=lambda x: x["dt"], reverse=True)
    return out


def _load_impulse() -> list:
    """Импульс памп (paper) — join signals(вход) + outcomes(выход) по signal_id.

    Метрика движка — net_pct (edge = ride-the-move, не TP/SL).
    outcome: sl / time / ride_exit (механизм выхода, не win/loss).
    win = net_pct > 0.
    """
    signals: dict = {}
    for r in _jsonl_multi("impulse_pump/impulse_pump_signals.jsonl"):
        sid = r.get("signal_id", "")
        if sid and sid not in signals:
            signals[sid] = r

    out = []
    seen: set = set()
    for rec in _jsonl_multi("impulse_pump/impulse_pump_outcomes.jsonl"):
        sid = rec.get("signal_id", "")
        if not sid or sid in seen:
            continue
        seen.add(sid)
        sig = signals.get(sid, {})
        ts  = sig.get("ts") or rec.get("ts") or ""
        out.append({
            "dt":          ts[5:16].replace("T", " ") if ts else "",
            "pair":        rec.get("symbol", "").replace("-USDT-SWAP", "").replace("-SWAP", ""),
            "side":        "LONG" if rec.get("side") == "long" else "SHORT",
            "entry":       sig.get("entry_price"),
            "exit":        rec.get("exit_price"),
            "outcome":     (rec.get("outcome") or "").upper(),
            "net_pct":     rec.get("net_pct"),
            "mfe_pct":     rec.get("mfe_pct"),
            "mae_pct":     rec.get("mae_pct"),
            "capture_pct": rec.get("capture_pct"),
            "hold_min":    rec.get("hold_min"),
            "valid":       rec.get("valid", True),
            "exit_rule":   sig.get("exit_rule") or rec.get("exit_rule"),
        })
    out.sort(key=lambda x: x["dt"], reverse=True)
    return out


def _load_main_impulse() -> list:
    signals: dict = {}
    for r in _jsonl_multi("main_impulse/main_impulse_signals.jsonl"):
        sid = r.get("signal_id", "")
        if sid and sid not in signals:
            signals[sid] = r

    out = []
    seen: set = set()
    for rec in _jsonl_multi("main_impulse/main_impulse_outcomes.jsonl"):
        sid = rec.get("signal_id", "")
        if not sid or sid in seen:
            continue
        seen.add(sid)
        sig = signals.get(sid, {})
        ts = sig.get("ts") or rec.get("ts") or ""
        out.append({
            "dt":          ts[5:16].replace("T", " ") if ts else "",
            "pair":        rec.get("symbol", "").replace("-USDT-SWAP", "").replace("-SWAP", ""),
            "side":        "LONG" if rec.get("side") == "long" else "SHORT",
            "entry":       sig.get("entry"),
            "exit":        rec.get("exit_price"),
            "outcome":     (rec.get("outcome") or "").upper(),
            "net_pct":     rec.get("net_pct"),
            "mfe_pct":     rec.get("mfe_pct"),
            "mae_pct":     rec.get("mae_pct"),
            "capture_pct": rec.get("capture_pct"),
            "hold_min":    rec.get("hold_min"),
            "valid":       rec.get("valid", True),
            "exit_rule":   sig.get("exit_rule") or rec.get("exit_rule"),
        })
    out.sort(key=lambda x: x["dt"], reverse=True)
    return out


def _load_bb_fade() -> list:
    """BB Fade signals — outcome пока приходит из самого signal jsonl (поле outcome)."""
    out = []
    for sig in _jsonl_multi("bb_fade/bb_fade_signals.jsonl"):
        ts = sig.get("ts", "")
        side = "LONG" if sig.get("side") == "buy" else "SHORT"
        outcome_raw = sig.get("outcome") or ""
        outcome = _outcome_to_tp_unified(outcome_raw)
        exit_r = _exit_r_at_tp1(sig.get("entry"), sig.get("sl"), sig.get("tp"), side, outcome_raw)
        out.append({
            "dt":         ts[5:16].replace("T", " ") if ts else "",
            "pair":       sig.get("symbol", "").replace("-USDT-SWAP", ""),
            "regime":     "BB_FADE",
            "style":      "FADE",
            "side":       side,
            "entry":      sig.get("entry"),
            "sl":         sig.get("sl"),
            "tp":         sig.get("tp"),
            "hold_min":   sig.get("hold_min"),
            "outcome":    outcome,
            "tp2_reached": False,
            "exit_price": None,
            "exit_r":     exit_r,
        })
    out.sort(key=lambda x: x["dt"], reverse=True)
    return out


def _load_all_simulator_signals() -> list:
    """Объединённый поток для Симулятора — все источники активных каналов с source-меткой."""
    rows: list = []
    for r in _load_main_ws():
        r2 = dict(r)
        r2["source"] = "Main WS"
        rows.append(r2)
    for r in _load_bb_fade():
        r2 = dict(r)
        r2["source"] = "BB Fade"
        rows.append(r2)
    rows.sort(key=lambda x: x["dt"], reverse=True)
    return rows


async def _fetch_positions_async() -> list:
    if os.getenv("JOURNAL_ENABLE_PRIVATE_FILLS", "").strip().lower() not in {"1", "true", "yes", "on"}:
        print("Ручные: private OKX fills skipped (set JOURNAL_ENABLE_PRIVATE_FILLS=1 to opt in)")
        return []
    api_key    = os.getenv("OKX_API_KEY", "")
    secret_key = os.getenv("OKX_SECRET_KEY", "")
    passphrase = os.getenv("OKX_PASSPHRASE", "")
    is_demo    = os.getenv("OKX_IS_DEMO", "1") == "1"
    if not api_key:
        return []
    client = _create_okx_client(api_key, secret_key, passphrase, is_demo=is_demo)
    try:
        inst_ids: set = set()
        after = None
        for _ in range(20):
            params: dict = {"instType": "SWAP", "limit": "100"}
            if after:
                params["after"] = after
            data  = await client._get("/api/v5/trade/fills-history", params)
            batch = data.get("data") or []
            for f in batch:
                iid = f.get("instId", "")
                if iid.endswith("-SWAP"):
                    inst_ids.add(iid)
            if len(batch) < 100:
                break
            after = batch[-1]["tradeId"]

        positions: list = []
        for iid in sorted(inst_ids):
            data = await client._get(
                "/api/v5/account/positions-history",
                {"instId": iid, "limit": "100"},
            )
            positions.extend(data.get("data") or [])
        return positions
    finally:
        await client.close()


def _create_okx_client(api_key: str, secret_key: str, passphrase: str, *, is_demo: bool):
    from src.exchange.okx_client import OKXClient

    return OKXClient(api_key, secret_key, passphrase, is_demo=is_demo)


def _load_real() -> list:
    try:
        positions = asyncio.run(_fetch_positions_async())
    except Exception as e:
        print(f"Ошибка загрузки реальных сделок: {e}")
        positions = []

    out = []
    for pos in positions:
        open_ms  = int(pos.get("cTime") or 0)
        close_ms = int(pos.get("uTime") or 0)
        symbol   = pos.get("instId", "").replace("-USDT-SWAP", "").replace("-SWAP", "")
        realized = float(pos.get("realizedPnl") or 0)
        funding  = float(pos.get("fundingFee")  or 0)
        net_pnl  = realized + funding
        pnl_pct  = float(pos.get("pnlRatio") or 0) * 100
        lever    = int(float(pos.get("lever") or 1))
        hold_min = round((close_ms - open_ms) / 60_000) if close_ms > open_ms else 0
        out.append({
            "dt":       _fmt_dt(close_ms),
            "pair":     symbol,
            "side":     (pos.get("direction") or "long").upper(),
            "entry":    float(pos.get("openAvgPx")  or 0) or None,
            "exit":     float(pos.get("closeAvgPx") or 0) or None,
            "lever":    lever,
            "net_pnl":  round(net_pnl, 4),
            "pnl_pct":  round(pnl_pct, 2),
            "hold_min": hold_min,
            "open_ms":  open_ms,
        })
    out.sort(key=lambda x: x["open_ms"], reverse=True)
    n_win = sum(1 for r in out if r["net_pnl"] > 0)
    print(f"Ручные: {len(out)} позиций, {n_win} прибыльных, avg win {sum(r['hold_min'] for r in out if r['net_pnl']>0)/n_win:.0f} мин" if n_win else f"Ручные: {len(out)} позиций")
    return out


# ── Sheet 1: Скринер ─────────────────────────────────────────────────────────
# A=date B=pair C=style D=regime E=side F=entry G=sl H=tp I=hold_min
# J=outcome K=R L=MFE_R M=MAE_R N=ADX_1H O=Slope_1H P=Slope_15m

def _build_screener(wb, rows: list) -> None:
    ws = wb.active
    ws.title = "Скринер"
    _hdr(ws, 1, [
        "Дата", "Пара", "Стиль", "Режим", "Сторона",
        "Вход", "SL", "TP", "Hold(мин)", "Исход",
        "R", "MFE_R", "MAE_R", "ADX_1H", "Slope_1H°", "Slope_15m°",
    ])
    for r, row in enumerate(rows, 2):
        vals = [
            row["dt"], row["pair"], row["style"], row["regime"], row["side"],
            row["entry"], row["sl"], row["tp"], row["hold_min"], row["outcome"],
            row["R"], row["MFE_R"], row["MAE_R"],
            round(row["adx_1h"], 1)   if row["adx_1h"]   is not None else "",
            round(row["slope_1h"], 1) if row["slope_1h"]  is not None else "",
            round(row["slope_15m"], 1) if row["slope_15m"] is not None else "",
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 10 and row["outcome"] in OUTCOME_FILL:
                cell.fill = OUTCOME_FILL[row["outcome"]]

    # Summary block using Excel formulas — live, always up to date
    sr = len(rows) + 3
    summary = [
        ("Всего сигналов", "=COUNTA(A2:A10000)"),
        ("TP",             '=COUNTIF(J2:J10000,"TP")'),
        ("SL",             '=COUNTIF(J2:J10000,"SL")'),
        ("TIME_EXIT",      '=COUNTIF(J2:J10000,"TIME_EXIT")'),
        ("WR% (TP vs SL)", '=IF(COUNTIF(J2:J10000,"TP")+COUNTIF(J2:J10000,"SL")=0,"",'
                           'ROUND(COUNTIF(J2:J10000,"TP")/(COUNTIF(J2:J10000,"TP")+COUNTIF(J2:J10000,"SL"))*100,1))'),
        ("Avg R",          '=IF(COUNTA(K2:K10000)=0,"",ROUND(AVERAGE(K2:K10000),2))'),
    ]
    for i, (lbl, val) in enumerate(summary):
        ws.cell(row=sr + i, column=1, value=lbl).fill = F_SUM
        ws.cell(row=sr + i, column=1).font = FONT_BOLD
        ws.cell(row=sr + i, column=2, value=val).fill = F_SUM

    _set_widths(ws, [13, 15, 7, 10, 8, 10, 10, 10, 9, 11, 7, 7, 7, 8, 10, 10])
    ws.freeze_panes = "A2"


# ── Sheet 2: Симулятор ───────────────────────────────────────────────────────
# A=Капитал$ B=Плечо C=date D=pair E=style F=side
# G=entry H=sl I=tp J=outcome K=R L=MAE_R
# M=Позиция$ N=Цена ликв O=Реал.исход P=P&L$ Q=P&L% R=Комис$ S=Финанс$ T=Чист P&L$

def _build_simulator(wb, rows: list) -> None:
    """Симулятор — Main WS + BB Fade с R-multiple клиента (TP1-only unification).

    Колонки A=капитал B=плечо — input (можно менять для каждой строки).
    P&L$ = position × (SL_dist_pct) × R. Где R — R-multiple до TP1 (клиент закрывает на TP1).
    """
    ws = wb.create_sheet("Симулятор")
    _hdr(ws, 1, [
        "Капитал$", "Плечо",          # A B — input
        "Дата", "Источник", "Пара", "Режим", "Стиль", "Сторона",  # C-H — info
        "Вход", "SL", "TP", "Исход", "R",                # I-M — signal data
        "Позиция$", "Цена ликв.", "Реал. исход",          # N-P — derived
        "P&L$", "P&L%", "Комиссия$", "Чист. P&L$",       # Q-T — money
    ])
    for c in (1, 2):
        ws.cell(row=1, column=c).fill = _fill("F4B942")

    for r, row in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=100).fill = F_INPUT
        ws.cell(row=r, column=2, value=10).fill  = F_INPUT

        data = {
            3:  row["dt"],
            4:  row.get("source", ""),
            5:  row["pair"],
            6:  row.get("regime", ""),
            7:  row.get("style", ""),
            8:  row["side"],
            9:  row["entry"],
            10: row["sl"],
            11: row["tp"],
            12: row["outcome"] or "",
            13: row.get("exit_r"),
        }
        for c, val in data.items():
            cell = ws.cell(row=r, column=c, value=val)
            if c == 12 and row["outcome"] in OUTCOME_FILL:
                cell.fill = OUTCOME_FILL[row["outcome"]]

        if not (row["entry"] and row["sl"] and row["tp"]):
            continue

        # N = Позиция$ = капитал × плечо
        ws.cell(row=r, column=14, value=f"=A{r}*B{r}")
        # O = Цена ликвидации (~maint margin 0.55%)
        ws.cell(row=r, column=15, value=(
            f'=IF(H{r}="LONG",I{r}*(1-1/B{r}+0.0055),I{r}*(1+1/B{r}-0.0055))'
        ))
        # P = Реал. исход (учитывает ликвидацию если до неё дошло)
        ws.cell(row=r, column=16, value=(
            f'=IF(L{r}="","",IF(AND(M{r}<>"",H{r}="LONG",'
            f'I{r}-M{r}*ABS(I{r}-J{r})<=O{r}),"ЛИКВИДАЦИЯ",'
            f'IF(AND(M{r}<>"",H{r}="SHORT",'
            f'I{r}+M{r}*ABS(J{r}-I{r})>=O{r}),"ЛИКВИДАЦИЯ",L{r})))'
        ))
        # Q = P&L$ = position × SL_dist_pct × R (или -капитал при ликвидации)
        ws.cell(row=r, column=17, value=(
            f'=IF(OR(P{r}="",M{r}=""),"",IF(P{r}="ЛИКВИДАЦИЯ",-A{r},'
            f'N{r}*ABS(I{r}-J{r})/I{r}*M{r}))'
        ))
        # R = P&L%
        ws.cell(row=r, column=18, value=(
            f'=IF(OR(Q{r}="",A{r}=0),"",Q{r}/A{r}*100)'
        ))
        # S = Комиссия$ = -0.1% от позиции (round-trip 0.05% × 2)
        ws.cell(row=r, column=19, value=(
            f'=IF(N{r}="","",-N{r}*0.001)'
        ))
        # T = Чистая P&L$ = P&L + комиссия
        ws.cell(row=r, column=20, value=(
            f'=IF(Q{r}="","",Q{r}+S{r})'
        ))

    # Summary block в конце
    sr = len(rows) + 3
    summary = [
        ("Всего сделок (с выходом)",  f'=COUNTA(L2:L{sr-2})-COUNTBLANK(L2:L{sr-2})'),
        ("Net P&L$ (чистый)",          f'=ROUND(SUM(T2:T{sr-2}),2)'),
        ("Прибыль$",                    f'=ROUND(SUMIF(T2:T{sr-2},">0",T2:T{sr-2}),2)'),
        ("Убыток$",                     f'=ROUND(SUMIF(T2:T{sr-2},"<0",T2:T{sr-2}),2)'),
        ("Profit Factor",               f'=IFERROR(SUMIF(T2:T{sr-2},">0",T2:T{sr-2})/ABS(SUMIF(T2:T{sr-2},"<0",T2:T{sr-2})),"—")'),
        ("ROI на капитал (avg %)",      f'=IFERROR(ROUND(AVERAGE(R2:R{sr-2}),2),"—")'),
    ]
    for i, (lbl, val) in enumerate(summary):
        ws.cell(row=sr + i, column=1, value=lbl).fill = F_SUM
        ws.cell(row=sr + i, column=1).font = FONT_BOLD
        ws.cell(row=sr + i, column=2, value=val).fill = F_SUM

    _set_widths(ws, [9, 6, 13, 10, 12, 10, 7, 8, 10, 10, 10, 8, 7, 11, 12, 13, 10, 8, 10, 12])
    ws.freeze_panes = "C2"
    print(f"Симулятор: {len(rows)} строк (Main WS + BB Fade)")


# ── Sheet 3: Main WS ─────────────────────────────────────────────────────────
# A=date B=pair C=regime D=side E=entry F=sl G=tp1 H=tp2 I=hold_min
# J=outcome K=exit_price

def _build_main_ws(wb, rows: list) -> None:
    """Main WS лист — унификация TP1/TP2 → TP (клиент торгует один TP)."""
    ws = wb.create_sheet("Main WS")
    _hdr(ws, 1, [
        "Дата", "Пара", "Режим", "Стиль", "Сторона",
        "Вход", "SL", "TP", "Hold(мин)",
        "Исход", "R", "Дошёл до TP2?", "Выход",
    ])
    for r, row in enumerate(rows, 2):
        vals = [
            row["dt"], row["pair"], row["regime"], row.get("style", ""), row["side"],
            row["entry"], row["sl"], row["tp"], row["hold_min"],
            row["outcome"], row.get("exit_r"),
            "✓" if row.get("tp2_reached") else "",
            row["exit_price"],
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 10 and row["outcome"] in OUTCOME_FILL:
                cell.fill = OUTCOME_FILL[row["outcome"]]

    labeled   = [r for r in rows if r["outcome"] and r["outcome"] != "INVALID"]
    n_invalid = sum(1 for r in rows if r["outcome"] == "INVALID")
    n_tp     = sum(1 for r in labeled if r["outcome"] == "TP")
    n_sl     = sum(1 for r in labeled if r["outcome"] == "SL")
    n_time   = sum(1 for r in labeled if r["outcome"] == "TIME")
    n_tp2    = sum(1 for r in labeled if r.get("tp2_reached"))
    decisive = n_tp + n_sl
    per_sig  = n_tp + n_sl + n_time  # TIME now a real outcome -> per-signal denominator
    wr_str   = f"{n_tp}/{decisive} = {n_tp/decisive*100:.0f}%" if decisive else "—"
    wr_psig  = f"{n_tp}/{per_sig} = {n_tp/per_sig*100:.0f}%" if per_sig else "—"
    net_vals   = [r["exit_r"] for r in labeled if r.get("exit_r") is not None]
    gross_vals = [r["exit_r_gross"] for r in labeled if r.get("exit_r_gross") is not None]
    sum_r_net    = round(sum(net_vals), 2) if net_vals else 0
    avg_r_net    = round(sum_r_net / len(net_vals), 3) if net_vals else 0
    sum_r_gross  = round(sum(gross_vals), 2) if gross_vals else 0
    avg_r_gross  = round(sum_r_gross / len(gross_vals), 3) if gross_vals else 0

    sr = len(rows) + 3
    for i, (lbl, val) in enumerate([
        ("Всего сигналов",        len(rows)),
        ("Лейблировано",          len(labeled)),
        ("INVALID (дегенерат.)",  n_invalid),
        ("WR (decisive, vs SL)",  wr_str),
        ("WR (per-signal, вкл.TIME)", wr_psig),
        ("TP (TP1+ дошедшие до TP2)", n_tp),
        ("SL",                    n_sl),
        ("TIME (вкл. в R)",       n_time),
        ("Дошло до TP2 (инфо)",   f"{n_tp2}/{n_tp}" if n_tp else "0"),
        ("Сумма R (GROSS)",       sum_r_gross),
        ("Сумма R (NET, с костами)", sum_r_net),
        ("Avg R (GROSS)",         avg_r_gross),
        ("Avg R (NET)",           avg_r_net),
    ]):
        ws.cell(row=sr + i, column=1, value=lbl).fill = F_SUM
        ws.cell(row=sr + i, column=1).font = FONT_BOLD
        ws.cell(row=sr + i, column=2, value=val).fill = F_SUM

    _set_widths(ws, [13, 14, 10, 8, 8, 10, 10, 10, 9, 8, 7, 12, 10])
    ws.freeze_panes = "A2"
    print(f"Main WS: {len(rows)} сигналов, {len(labeled)} лейбл. (INVALID {n_invalid}), "
          f"WR decisive={wr_str}, WR per-signal={wr_psig}, sumR NET={sum_r_net} (GROSS {sum_r_gross})")


# ── Sheet 3b: BB Fade ─────────────────────────────────────────────────────────

def _build_bb_fade(wb, rows: list) -> None:
    """BB Fade лист — отдельный канал mean reversion. TP единственный (нет TP1/TP2)."""
    ws = wb.create_sheet("BB Fade")
    _hdr(ws, 1, [
        "Дата", "Пара", "Сторона", "Вход", "SL", "TP",
        "Hold(мин)", "Исход", "R",
    ])
    for r, row in enumerate(rows, 2):
        vals = [
            row["dt"], row["pair"], row["side"], row["entry"], row["sl"], row["tp"],
            row["hold_min"], row["outcome"], row.get("exit_r"),
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 8 and row["outcome"] in OUTCOME_FILL:
                cell.fill = OUTCOME_FILL[row["outcome"]]

    if not rows:
        ws.cell(row=2, column=1, value="Нет данных пока (ждём первые сигналы)")

    labeled  = [r for r in rows if r["outcome"]]
    n_tp     = sum(1 for r in labeled if r["outcome"] == "TP")
    n_sl     = sum(1 for r in labeled if r["outcome"] == "SL")
    decisive = n_tp + n_sl
    wr_str   = f"{n_tp}/{decisive} = {n_tp/decisive*100:.0f}%" if decisive else "—"

    sr = max(len(rows), 1) + 3
    for i, (lbl, val) in enumerate([
        ("Всего сигналов",     len(rows)),
        ("Лейблировано",       len(labeled)),
        ("WR (decisive)",      wr_str),
    ]):
        ws.cell(row=sr + i, column=1, value=lbl).fill = F_SUM
        ws.cell(row=sr + i, column=1).font = FONT_BOLD
        ws.cell(row=sr + i, column=2, value=val).fill = F_SUM

    _set_widths(ws, [13, 14, 8, 10, 10, 10, 9, 8, 7])
    ws.freeze_panes = "A2"
    print(f"BB Fade: {len(rows)} сигналов, WR={wr_str}")


# ── Sheet 4: Памп ────────────────────────────────────────────────────────────
# A=date B=pair C=entry D=sl E=tp F=outcome G=hold_min H=signal_hour I=net_pnl_pct

_PUMP_V2_DATE = "10.05"  # first date of new orchestrator

def _build_pump(wb, rows: list) -> None:
    ws = wb.create_sheet("Памп")
    _hdr(ws, 1, [
        "Дата", "Пара", "Вход", "SL", "TP",
        "Исход", "Hold(мин)", "Час UTC", "NET P&L%",
    ])

    excel_row = 2
    separator_written = False
    for row in rows:
        # Insert separator between old engine and new orchestrator
        day = row["dt"][:5] if row["dt"] else ""
        if not separator_written and day and day < _PUMP_V2_DATE:
            sep_fill = _fill("FFD966")  # gold
            sep_font = Font(bold=True, color="7F4F00")
            for c in range(1, 10):
                cell = ws.cell(row=excel_row, column=c)
                cell.fill = sep_fill
                cell.font = sep_font
            ws.cell(row=excel_row, column=1,
                    value="▲ Новый оркестратор (с 10.05.2026) ▲").font = sep_font
            ws.cell(row=excel_row, column=1).fill = sep_fill
            ws.merge_cells(start_row=excel_row, start_column=1,
                           end_row=excel_row, end_column=9)
            ws.cell(row=excel_row, column=1).alignment = ALIGN_C
            excel_row += 1
            separator_written = True

        vals = [
            row["dt"], row["pair"], row["entry"], row["sl"], row["tp"],
            row["outcome"], row["hold"], row["hour"], row["net_pnl"],
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=excel_row, column=c, value=val)
            if c == 6 and row["outcome"] in OUTCOME_FILL:
                cell.fill = OUTCOME_FILL[row["outcome"]]
        excel_row += 1

    n = len(rows)
    if n:
        n_tp   = sum(1 for r in rows if r["outcome"] == "TP")
        wr     = n_tp / n * 100
        wins   = [r["net_pnl"] for r in rows if r["outcome"] == "TP"  and r["net_pnl"] is not None]
        losses = [abs(r["net_pnl"]) for r in rows if r["outcome"] == "SL" and r["net_pnl"] is not None]
        pf     = sum(wins) / sum(losses) if losses else 999.0
        avg    = sum(r["net_pnl"] for r in rows if r["net_pnl"] is not None) / n

        # +1 for separator row if it was written
        sr = excel_row + 2
        for i, (lbl, val) in enumerate([
            ("Всего сделок",  n),
            ("WR",            f"{n_tp}/{n} = {wr:.0f}%"),
            ("Profit Factor", f"{pf:.2f}"),
            ("Avg NET P&L%",  f"{avg:+.3f}%"),
        ]):
            ws.cell(row=sr + i, column=1, value=lbl).fill = F_SUM
            ws.cell(row=sr + i, column=1).font = FONT_BOLD
            ws.cell(row=sr + i, column=2, value=val).fill = F_SUM

    _set_widths(ws, [13, 12, 11, 11, 11, 8, 9, 8, 10])
    ws.freeze_panes = "A2"
    print(f"Памп: {n} сделок")


# ── Sheet 4b: Импульс ────────────────────────────────────────────────────────
# A=date B=pair C=side D=entry E=exit F=outcome G=NET% H=MFE% I=MAE% J=Cap% K=hold L=valid
# Метрика — net_pct (ride-the-move). outcome=sl/time/ride_exit — механизм выхода.

_IMPULSE_FILL = {"SL": F_SL, "TIME": F_TIME, "RIDE_EXIT": F_TP, "SCALED_TP": F_TP}


def _build_impulse(wb, rows: list, sheet_name: str = "Импульс") -> None:
    """Импульс памп (paper) — рывок-вход + ride. WR по net_pct>0, не по TP/SL."""
    ws = wb.create_sheet(sheet_name)
    _hdr(ws, 1, [
        "Дата", "Пара", "Сторона", "Вход", "Выход",
        "Выход(тип)", "NET %", "MFE %", "MAE %", "Capture %",
        "Hold(мин)", "Exit rule", "Valid",
    ])

    def _r(v, n=3):
        return round(v, n) if isinstance(v, (int, float)) else v

    for r, row in enumerate(rows, 2):
        vals = [
            row["dt"], row["pair"], row["side"], row["entry"], row["exit"],
            row["outcome"], _r(row["net_pct"]), _r(row["mfe_pct"]), _r(row["mae_pct"]),
            _r(row["capture_pct"], 1), _r(row["hold_min"], 1),
            (row.get("exit_rule") or {}).get("type") if isinstance(row.get("exit_rule"), dict) else "",
            "" if row["valid"] else "✗",
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 6 and row["outcome"] in _IMPULSE_FILL:
                cell.fill = _IMPULSE_FILL[row["outcome"]]
            if c == 7 and isinstance(row["net_pct"], (int, float)):
                cell.fill = F_TP if row["net_pct"] >= 0 else F_SL

    if not rows:
        ws.cell(row=2, column=1, value="Нет данных пока (paper не запущен / ждём первые сделки)")

    closed = [r for r in rows if isinstance(r["net_pct"], (int, float))]
    n      = len(closed)
    if n:
        wins   = [r["net_pct"] for r in closed if r["net_pct"] > 0]
        losses = [r["net_pct"] for r in closed if r["net_pct"] < 0]
        wr     = len(wins) / n * 100
        pf     = sum(wins) / abs(sum(losses)) if losses else 999.0
        avg    = sum(r["net_pct"] for r in closed) / n
        net    = sum(r["net_pct"] for r in closed)
        caps   = [r["capture_pct"] for r in closed
                  if isinstance(r["capture_pct"], (int, float))]
        avg_cap = sum(caps) / len(caps) if caps else 0.0
        holds   = [r["hold_min"] for r in closed
                   if isinstance(r["hold_min"], (int, float))]
        avg_hold = sum(holds) / len(holds) if holds else 0.0

        sr = len(rows) + 3
        for i, (lbl, val) in enumerate([
            ("Всего сделок (закрыто)", n),
            ("WR (net>0)",            f"{len(wins)}/{n} = {wr:.0f}%"),
            ("Profit Factor",         f"{pf:.2f}"),
            ("Avg NET %",             f"{avg:+.3f}%"),
            ("Сумма NET %",           f"{net:+.2f}%"),
            ("Avg Capture %",         f"{avg_cap:.0f}%"),
            ("Avg Hold (мин)",        f"{avg_hold:.0f}"),
        ]):
            ws.cell(row=sr + i, column=1, value=lbl).fill = F_SUM
            ws.cell(row=sr + i, column=1).font = FONT_BOLD
            ws.cell(row=sr + i, column=2, value=val).fill = F_SUM

    _set_widths(ws, [13, 12, 8, 11, 11, 11, 9, 8, 8, 10, 9, 11, 6])
    ws.freeze_panes = "A2"
    print(f"{sheet_name}: {len(rows)} сделок ({len(closed)} закрыто)")


# ── Sheet 5: Ручные ──────────────────────────────────────────────────────────
# A=date B=pair C=side D=entry E=exit F=lever G=net_pnl$ H=pnl_pct% I=hold_min

def _build_real(wb, rows: list) -> None:
    ws = wb.create_sheet("Ручные")
    _hdr(ws, 1, [
        "Дата закр.", "Пара", "Сторона",
        "Вход", "Выход", "Плечо",
        "NET P&L$", "P&L%", "Hold(мин)",
    ])
    for r, row in enumerate(rows, 2):
        vals = [
            row["dt"], row["pair"], row["side"],
            row["entry"], row["exit"], row["lever"],
            row["net_pnl"], row["pnl_pct"], row["hold_min"],
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 7:
                cell.fill = F_TP if (row["net_pnl"] or 0) >= 0 else F_SL

    if not rows:
        ws.cell(row=2, column=1, value="Нет данных (OKX не вернул позиции)")

    # Summary block — Excel formulas, column G=net_pnl$, I=hold_min
    sr = max(len(rows), 1) + 3
    summary = [
        ("Всего позиций",   "=COUNTA(A2:A10000)"),
        ("Побед",           '=COUNTIF(G2:G10000,">0")'),
        ("Потерь",          '=COUNTIF(G2:G10000,"<0")'),
        ("NET P&L$",        "=ROUND(SUM(G2:G10000),2)"),
        ("Прибыль$",        '=ROUND(SUMIF(G2:G10000,">0",G2:G10000),2)'),
        ("Убыток$",         '=ROUND(SUMIF(G2:G10000,"<0",G2:G10000),2)'),
        ("Avg hold WIN(мин)",  '=ROUND(AVERAGEIF(G2:G10000,">0",I2:I10000),0)'),
        ("Avg hold LOSS(мин)", '=ROUND(AVERAGEIF(G2:G10000,"<0",I2:I10000),0)'),
    ]
    for i, (lbl, val) in enumerate(summary):
        ws.cell(row=sr + i, column=1, value=lbl).fill = F_SUM
        ws.cell(row=sr + i, column=1).font = FONT_BOLD
        ws.cell(row=sr + i, column=2, value=val).fill = F_SUM

    _set_widths(ws, [13, 12, 8, 11, 11, 7, 11, 8, 10])
    ws.freeze_panes = "A2"


# ── Sheet 6: Дашборд ─────────────────────────────────────────────────────────

def _build_dashboard(wb) -> None:
    ws = wb.create_sheet("Дашборд")
    ws.sheet_view.showGridLines = False

    # ── Блок A: Итоги по системам (A1:F7) ────────────────────────────────────
    # WR DECISIVE: TP/(TP+SL), TIME вынесен в отдельную колонку справочно
    _dash_title(ws, 1, 1, "ИТОГИ ПО СИСТЕМАМ (WR decisive = TP/TP+SL)", span=6)
    _dash_hdr(ws, 2, 1, ["Система", "TP", "SL", "TIME", "Всего", "WR%"])

    systems = [
        ("Скринер",
         '=COUNTIF(Скринер!J2:J10000,"TP")+COUNTIF(Скринер!J2:J10000,"TP1")+COUNTIF(Скринер!J2:J10000,"TP2")',
         '=COUNTIF(Скринер!J2:J10000,"SL")',
         '=COUNTIF(Скринер!J2:J10000,"TIME")+COUNTIF(Скринер!J2:J10000,"TIME_EXIT")',
         ),
        ("Main WS",
         '=COUNTIF(\'Main WS\'!J2:J10000,"TP")',  # унифицировано: TP1+TP2 → TP в loader
         '=COUNTIF(\'Main WS\'!J2:J10000,"SL")',
         '=COUNTIF(\'Main WS\'!J2:J10000,"TIME")',
         ),
        ("BB Fade",
         '=COUNTIF(\'BB Fade\'!H2:H10000,"TP")',
         '=COUNTIF(\'BB Fade\'!H2:H10000,"SL")',
         '=COUNTIF(\'BB Fade\'!H2:H10000,"TIME")',
         ),
        ("Памп",
         '=COUNTIF(Памп!F2:F10000,"TP")',
         '=COUNTIF(Памп!F2:F10000,"SL")',
         '=COUNTIF(Памп!F2:F10000,"TIME")',
         ),
        ("Импульс",  # win=net>0 (G), loss=net<0; не TP/SL
         '=COUNTIF(Импульс!G2:G10000,">0")',
         '=COUNTIF(Импульс!G2:G10000,"<0")',
         "0",
         ),
        ("Ручные",
         '=COUNTIF(Ручные!G2:G10000,">0")',
         '=COUNTIF(Ручные!G2:G10000,"<0")',
         "0",
         ),
    ]
    for i, (name, tp_f, sl_f, t_f) in enumerate(systems):
        r    = 3 + i
        fill = F_STRIPE if i % 2 == 0 else None
        _dw(ws, r, 1, name, fill=fill, font=FONT_BOLD)
        _dw(ws, r, 2, tp_f,  fill=fill)
        _dw(ws, r, 3, sl_f,  fill=fill)
        _dw(ws, r, 4, t_f,   fill=fill)
        _dw(ws, r, 5, f"=B{r}+C{r}+D{r}", fill=fill)
        # WR decisive: TP/(TP+SL)
        _dw(ws, r, 6, f'=IF(B{r}+C{r}=0,"",ROUND(B{r}/(B{r}+C{r})*100,1))', fill=fill)

    # ── Блок B: Скринер по стилю (A10:F14) ────────────────────────────────────
    _dash_title(ws, 10, 1, "СКРИНЕР — ПО СТИЛЮ", span=6)
    _dash_hdr(ws, 11, 1, ["Стиль", "TP", "SL", "TIME", "Всего", "WR%"])
    for i, style in enumerate(["FAST", "SWING", "FADE"]):
        r    = 12 + i
        fill = F_STRIPE if i % 2 == 0 else None
        _dw(ws, r, 1, style, fill=fill, font=FONT_BOLD)
        # TP catches "TP", "TP1", "TP2" — combined as wins
        _dw(ws, r, 2, f'=COUNTIFS(Скринер!C2:C10000,A{r},Скринер!J2:J10000,"TP")+'
                      f'COUNTIFS(Скринер!C2:C10000,A{r},Скринер!J2:J10000,"TP1")+'
                      f'COUNTIFS(Скринер!C2:C10000,A{r},Скринер!J2:J10000,"TP2")', fill=fill)
        _dw(ws, r, 3, f'=COUNTIFS(Скринер!C2:C10000,A{r},Скринер!J2:J10000,"SL")', fill=fill)
        _dw(ws, r, 4, f'=COUNTIFS(Скринер!C2:C10000,A{r},Скринер!J2:J10000,"TIME")+'
                      f'COUNTIFS(Скринер!C2:C10000,A{r},Скринер!J2:J10000,"TIME_EXIT")', fill=fill)
        _dw(ws, r, 5, f"=B{r}+C{r}+D{r}", fill=fill)
        _dw(ws, r, 6, f'=IF(B{r}+C{r}=0,"",ROUND(B{r}/(B{r}+C{r})*100,1))', fill=fill)

    # ── Блок C: Main WS по режиму (A16:F20) ──────────────────────────────────
    # Унификация: TP1+TP2 в loader уже схлопнуты в "TP", показываем "Дошло до TP2?" отдельно
    _dash_title(ws, 16, 1, "MAIN WS — ПО РЕЖИМУ (WR decisive)", span=6)
    _dash_hdr(ws, 17, 1, ["Режим", "TP", "SL", "TIME", "Всего", "WR%"])
    for i, regime in enumerate(["DRIFT", "TRENDING", "RANGING", "CHOPPY"]):
        r    = 18 + i
        fill = F_STRIPE if i % 2 == 0 else None
        _dw(ws, r, 1, regime, fill=fill, font=FONT_BOLD)
        _dw(ws, r, 2, f'=COUNTIFS(\'Main WS\'!C2:C10000,A{r},\'Main WS\'!J2:J10000,"TP")', fill=fill)
        _dw(ws, r, 3, f'=COUNTIFS(\'Main WS\'!C2:C10000,A{r},\'Main WS\'!J2:J10000,"SL")', fill=fill)
        _dw(ws, r, 4, f'=COUNTIFS(\'Main WS\'!C2:C10000,A{r},\'Main WS\'!J2:J10000,"TIME")', fill=fill)
        _dw(ws, r, 5, f"=B{r}+C{r}+D{r}", fill=fill)
        _dw(ws, r, 6, f'=IF(B{r}+C{r}=0,"",ROUND(B{r}/(B{r}+C{r})*100,1))', fill=fill)

    # ── Блок D: Main WS по стилю (H1:M6) ─────────────────────────────────────
    _dash_title(ws, 1, 8, "MAIN WS — ПО СТИЛЮ (WR decisive)", span=6)
    _dash_hdr(ws, 2, 8, ["Стиль", "TP", "SL", "TIME", "Всего", "WR%"])
    for i, style in enumerate(["FAST", "SWING", "BB_FADE"]):
        r    = 3 + i
        fill = F_STRIPE if i % 2 == 0 else None
        _dw(ws, r, 8,  style, fill=fill, font=FONT_BOLD)
        _dw(ws, r, 9,  f'=COUNTIFS(\'Main WS\'!D2:D10000,H{r},\'Main WS\'!J2:J10000,"TP")', fill=fill)
        _dw(ws, r, 10, f'=COUNTIFS(\'Main WS\'!D2:D10000,H{r},\'Main WS\'!J2:J10000,"SL")', fill=fill)
        _dw(ws, r, 11, f'=COUNTIFS(\'Main WS\'!D2:D10000,H{r},\'Main WS\'!J2:J10000,"TIME")', fill=fill)
        _dw(ws, r, 12, f"=I{r}+J{r}+K{r}", fill=fill)
        _dw(ws, r, 13, f'=IF(I{r}+J{r}=0,"",ROUND(I{r}/(I{r}+J{r})*100,1))', fill=fill)

    # ── Блок E: P&L сводка (H9:M15) ──────────────────────────────────────────
    _dash_title(ws, 9, 8, "P&L СВОДКА", span=6)
    _dash_hdr(ws, 10, 8, ["Источник", "Побед", "Потерь", "Net", "Прибыль", "Убыток"])
    pnl_data = [
        ("Main WS (R)",  # колонка K = exit_r в листе Main WS
         '=COUNTIF(\'Main WS\'!K2:K10000,">0")',
         '=COUNTIF(\'Main WS\'!K2:K10000,"<0")',
         '=ROUND(SUM(\'Main WS\'!K2:K10000),2)',
         '=ROUND(SUMIF(\'Main WS\'!K2:K10000,">0",\'Main WS\'!K2:K10000),2)',
         '=ROUND(SUMIF(\'Main WS\'!K2:K10000,"<0",\'Main WS\'!K2:K10000),2)',
         ),
        ("BB Fade (R)",  # колонка I = exit_r в листе BB Fade
         '=COUNTIF(\'BB Fade\'!I2:I10000,">0")',
         '=COUNTIF(\'BB Fade\'!I2:I10000,"<0")',
         '=ROUND(SUM(\'BB Fade\'!I2:I10000),2)',
         '=ROUND(SUMIF(\'BB Fade\'!I2:I10000,">0",\'BB Fade\'!I2:I10000),2)',
         '=ROUND(SUMIF(\'BB Fade\'!I2:I10000,"<0",\'BB Fade\'!I2:I10000),2)',
         ),
        ("Памп (%)",
         '=COUNTIF(Памп!I2:I10000,">0")',
         '=COUNTIF(Памп!I2:I10000,"<0")',
         '=ROUND(SUM(Памп!I2:I10000),2)',
         '=ROUND(SUMIF(Памп!I2:I10000,">0",Памп!I2:I10000),2)',
         '=ROUND(SUMIF(Памп!I2:I10000,"<0",Памп!I2:I10000),2)',
         ),
        ("Импульс (%)",  # колонка G = NET %
         '=COUNTIF(Импульс!G2:G10000,">0")',
         '=COUNTIF(Импульс!G2:G10000,"<0")',
         '=ROUND(SUM(Импульс!G2:G10000),2)',
         '=ROUND(SUMIF(Импульс!G2:G10000,">0",Импульс!G2:G10000),2)',
         '=ROUND(SUMIF(Импульс!G2:G10000,"<0",Импульс!G2:G10000),2)',
         ),
        ("Ручные ($)",
         '=COUNTIF(Ручные!G2:G10000,">0")',
         '=COUNTIF(Ручные!G2:G10000,"<0")',
         '=ROUND(SUM(Ручные!G2:G10000),2)',
         '=ROUND(SUMIF(Ручные!G2:G10000,">0",Ручные!G2:G10000),2)',
         '=ROUND(SUMIF(Ручные!G2:G10000,"<0",Ручные!G2:G10000),2)',
         ),
        ("Симулятор ($)",
         '=COUNTIF(Симулятор!T2:T10000,">0")',
         '=COUNTIF(Симулятор!T2:T10000,"<0")',
         '=ROUND(SUM(Симулятор!T2:T10000),2)',
         '=ROUND(SUMIF(Симулятор!T2:T10000,">0",Симулятор!T2:T10000),2)',
         '=ROUND(SUMIF(Симулятор!T2:T10000,"<0",Симулятор!T2:T10000),2)',
         ),
    ]
    for i, (src, wins, losses, net, profit, loss) in enumerate(pnl_data):
        r    = 11 + i
        fill = F_STRIPE if i % 2 == 0 else None
        _dw(ws, r, 8,  src,    fill=fill, font=FONT_BOLD)
        _dw(ws, r, 9,  wins,   fill=fill)
        _dw(ws, r, 10, losses, fill=fill)
        _dw(ws, r, 11, net,    fill=fill)
        _dw(ws, r, 12, profit, fill=_fill("C6EFCE"))
        _dw(ws, r, 13, loss,   fill=_fill("FFC7CE"))

    # ── Блок F: Памп по часам UTC (A24:F47) ──────────────────────────────────
    _dash_title(ws, 24, 1, "ПАМП — ПО ЧАСАМ UTC (WR decisive)", span=6)
    _dash_hdr(ws, 25, 1, ["Час", "TP", "SL", "Всего", "WR%", "NET P&L% сумм"])
    for hour in range(24):
        r    = 26 + hour
        fill = F_STRIPE if hour % 2 == 0 else None
        _dw(ws, r, 1, hour, fill=fill)
        _dw(ws, r, 2, f'=COUNTIFS(Памп!H2:H10000,A{r},Памп!F2:F10000,"TP")', fill=fill)
        _dw(ws, r, 3, f'=COUNTIFS(Памп!H2:H10000,A{r},Памп!F2:F10000,"SL")', fill=fill)
        _dw(ws, r, 4, f"=B{r}+C{r}", fill=fill)
        _dw(ws, r, 5, f'=IF(D{r}=0,"",ROUND(B{r}/D{r}*100,1))', fill=fill)
        _dw(ws, r, 6, f"=ROUND(SUMIF(Памп!H2:H10000,A{r},Памп!I2:I10000),2)", fill=fill)

    # ── ЧЕРТА: сброс под новую стратегию ──────────────────────────────────────
    _dash_title(ws, 51, 1,
                f"🔴 СБРОС {RESET_DATE} — все данные = СТАРАЯ стратегия (архив pre-reset_2026-05-27)", span=6)
    note = ws.cell(row=52, column=1,
                   value="Новый запуск пишет свежее в logs/. Новые сделки (новая стратегия) появятся сверху листов по мере накопления.")
    note.font = Font(italic=True, color="7F4F00")
    ws.merge_cells(start_row=52, start_column=1, end_row=52, end_column=6)

    _set_widths(ws, [12, 9, 9, 9, 8, 15, 3, 16, 9, 9, 11, 12, 12])


# ── Sheet 7: Графики ──────────────────────────────────────────────────────────

def _agg_by_day(rows: list, value_key: str) -> tuple[list, list, list]:
    """Группирует по дню → (days, daily_pnl, cumulative).

    Поддерживает оба формата dt:
      - "MM-DD HH:MM"          (Main WS, BB Fade) → берём [:5]
      - "YYYY-MM-DD HH:MM"     (Pump)             → берём [5:10] для MM-DD
    """
    by_day: dict[str, float] = {}
    for row in rows:
        dt = row.get("dt")
        if not dt:
            continue
        val = row.get(value_key)
        if val is None:
            continue
        try:
            val = float(val)
        except (TypeError, ValueError):
            continue
        # Нормализуем к MM-DD
        if len(dt) >= 10 and dt[4] == "-":
            day = dt[5:10]  # "2026-05-17 07:51" → "05-17"
        else:
            day = dt[:5]    # "05-17 06:15" → "05-17"
        by_day[day] = by_day.get(day, 0) + val
    days = sorted(by_day.keys())
    daily = [round(by_day[d], 2) for d in days]
    cum = []
    s = 0
    for v in daily:
        s += v
        cum.append(round(s, 2))
    return days, daily, cum


def _outcome_counts(rows: list, outcome_field: str = "outcome",
                    win_labels=("TP",), loss_labels=("SL",),
                    time_labels=("TIME",)) -> tuple[int, int, int]:
    win = sum(1 for r in rows if r.get(outcome_field) in win_labels)
    loss = sum(1 for r in rows if r.get(outcome_field) in loss_labels)
    time_ = sum(1 for r in rows if r.get(outcome_field) in time_labels)
    return win, loss, time_


def _build_charts(wb, main_ws_rows: list, bb_fade_rows: list, pump_rows: list) -> None:
    """Лист 'Графики' — visualizations per system. Equity curve + outcome pie."""
    ws = wb.create_sheet("Графики")
    ws.sheet_view.showGridLines = False

    # Data tables hidden in columns AA-AZ for chart sources
    col = 27  # AA
    row_start = 1

    # === MAIN WS ===
    ws.cell(row=1, column=1, value="📊 MAIN WS — основной канал").font = Font(bold=True, size=14, color="1F4E79")

    # Data: equity curve по дням (Net R)
    main_days, main_daily, main_cum = _agg_by_day(main_ws_rows, "exit_r")
    if main_days:
        ws.cell(row=row_start, column=col, value="День")
        ws.cell(row=row_start, column=col+1, value="Net R за день")
        ws.cell(row=row_start, column=col+2, value="Кумул. R")
        for i, (d, daily, cum) in enumerate(zip(main_days, main_daily, main_cum)):
            ws.cell(row=row_start+1+i, column=col, value=d)
            ws.cell(row=row_start+1+i, column=col+1, value=daily)
            ws.cell(row=row_start+1+i, column=col+2, value=cum)

        # LineChart кумулятивный
        chart = LineChart()
        chart.title = "Equity curve (Net R накопительно)"
        chart.style = 12
        chart.y_axis.title = "R"
        chart.x_axis.title = "Дата"
        chart.height = 8
        chart.width = 16
        data = Reference(ws, min_col=col+2, min_row=row_start, max_row=row_start+len(main_days), max_col=col+2)
        cats = Reference(ws, min_col=col, min_row=row_start+1, max_row=row_start+len(main_days))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "A3")

    # Outcome pie для Main WS
    main_labeled = [r for r in main_ws_rows if r.get("outcome")]
    n_tp, n_sl, n_time = _outcome_counts(main_labeled)
    if n_tp + n_sl + n_time > 0:
        pie_row = row_start + max(len(main_days), 5) + 3
        ws.cell(row=pie_row, column=col, value="Исход")
        ws.cell(row=pie_row, column=col+1, value="Количество")
        ws.cell(row=pie_row+1, column=col, value="TP")
        ws.cell(row=pie_row+1, column=col+1, value=n_tp)
        ws.cell(row=pie_row+2, column=col, value="SL")
        ws.cell(row=pie_row+2, column=col+1, value=n_sl)
        ws.cell(row=pie_row+3, column=col, value="TIME")
        ws.cell(row=pie_row+3, column=col+1, value=n_time)

        pie = PieChart()
        pie.title = f"Main WS: исходы (n={n_tp+n_sl+n_time})"
        labels = Reference(ws, min_col=col,   min_row=pie_row+1, max_row=pie_row+3)
        data   = Reference(ws, min_col=col+1, min_row=pie_row,   max_row=pie_row+3)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.dataLabels = DataLabelList(showPercent=True)
        pie.height = 8
        pie.width = 10
        ws.add_chart(pie, "K3")

    # === BB FADE ===
    ws.cell(row=20, column=1, value="📊 BB FADE — mean reversion").font = Font(bold=True, size=14, color="1F4E79")

    if bb_fade_rows:
        bb_days, bb_daily, bb_cum = _agg_by_day(bb_fade_rows, "exit_r")
        if bb_days:
            ws.cell(row=20, column=col, value="День")
            ws.cell(row=20, column=col+1, value="Net R за день")
            ws.cell(row=20, column=col+2, value="Кумул. R")
            for i, (d, daily, cum) in enumerate(zip(bb_days, bb_daily, bb_cum)):
                ws.cell(row=21+i, column=col,   value=d)
                ws.cell(row=21+i, column=col+1, value=daily)
                ws.cell(row=21+i, column=col+2, value=cum)
            chart = LineChart()
            chart.title = "BB Fade: equity curve (Net R)"
            chart.style = 12
            chart.height = 8
            chart.width = 16
            data = Reference(ws, min_col=col+2, min_row=20, max_row=20+len(bb_days), max_col=col+2)
            cats = Reference(ws, min_col=col,   min_row=21, max_row=20+len(bb_days))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "A22")
    else:
        ws.cell(row=22, column=1, value="(данных мало — 1 сигнал за 36ч live)").font = Font(italic=True, color="808080")

    # === PUMP ===
    ws.cell(row=40, column=1, value="📊 PUMP — alt scalp").font = Font(bold=True, size=14, color="1F4E79")

    pump_days, pump_daily, pump_cum = _agg_by_day(pump_rows, "net_pnl")
    if pump_days:
        ws.cell(row=40, column=col, value="День")
        ws.cell(row=40, column=col+1, value="Net % за день")
        ws.cell(row=40, column=col+2, value="Кумул. %")
        for i, (d, daily, cum) in enumerate(zip(pump_days, pump_daily, pump_cum)):
            ws.cell(row=41+i, column=col,   value=d)
            ws.cell(row=41+i, column=col+1, value=daily)
            ws.cell(row=41+i, column=col+2, value=cum)

        chart = LineChart()
        chart.title = "Pump: equity curve (Net %, накопительно)"
        chart.style = 12
        chart.height = 8
        chart.width = 16
        data = Reference(ws, min_col=col+2, min_row=40, max_row=40+len(pump_days), max_col=col+2)
        cats = Reference(ws, min_col=col,   min_row=41, max_row=40+len(pump_days))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "A42")

    # Outcome pie для Pump
    pump_tp = sum(1 for r in pump_rows if r.get("outcome") == "TP")
    pump_sl = sum(1 for r in pump_rows if r.get("outcome") == "SL")
    if pump_tp + pump_sl > 0:
        pie_row = 60
        ws.cell(row=pie_row, column=col,   value="Исход")
        ws.cell(row=pie_row, column=col+1, value="Количество")
        ws.cell(row=pie_row+1, column=col, value="TP")
        ws.cell(row=pie_row+1, column=col+1, value=pump_tp)
        ws.cell(row=pie_row+2, column=col, value="SL")
        ws.cell(row=pie_row+2, column=col+1, value=pump_sl)
        pie = PieChart()
        pie.title = f"Pump: исходы (n={pump_tp+pump_sl})"
        labels = Reference(ws, min_col=col,   min_row=pie_row+1, max_row=pie_row+2)
        data   = Reference(ws, min_col=col+1, min_row=pie_row,   max_row=pie_row+2)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.dataLabels = DataLabelList(showPercent=True)
        pie.height = 8
        pie.width = 10
        ws.add_chart(pie, "K42")

    # WR by hour bar chart (Pump)
    if pump_rows:
        from collections import defaultdict
        hour_stats = defaultdict(lambda: {"TP": 0, "SL": 0})
        for r in pump_rows:
            h = r.get("hour")
            o = r.get("outcome")
            if h is None or o not in ("TP", "SL"):
                continue
            hour_stats[h][o] += 1

        bar_row = 80
        ws.cell(row=bar_row, column=col,   value="Час UTC")
        ws.cell(row=bar_row, column=col+1, value="WR %")
        for hour in range(24):
            stat = hour_stats[hour]
            total = stat["TP"] + stat["SL"]
            wr = round(stat["TP"]/total*100, 1) if total else 0
            ws.cell(row=bar_row+1+hour, column=col,   value=hour)
            ws.cell(row=bar_row+1+hour, column=col+1, value=wr)
        bar = BarChart()
        bar.title = "Pump: WR % по часам UTC (decisive)"
        bar.y_axis.title = "WR %"
        bar.x_axis.title = "Час UTC"
        bar.style = 11
        bar.height = 8
        bar.width = 16
        data = Reference(ws, min_col=col+1, min_row=bar_row, max_row=bar_row+24)
        cats = Reference(ws, min_col=col,   min_row=bar_row+1, max_row=bar_row+24)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        ws.add_chart(bar, "A62")

    # Не скрываем data columns — некоторые Excel клиенты не рендерят графики
    # на скрытых столбцах. Просто узкие.
    for c in range(col, col+5):
        ws.column_dimensions[get_column_letter(c)].width = 10

    _set_widths(ws, [20] + [10]*9 + [4] + [20] + [10]*8)


def _build_paper_watch(wb, rows: list) -> None:
    """Paper-watch outcomes from the private Strategy Lab derived export."""
    ws = wb.create_sheet("Paper Watch")
    _hdr(ws, 1, [
        "Created", "Symbol", "OKX", "TF", "Family", "Side", "Exit mode", "Status",
        "Result", "Net %", "Net R", "MFE %", "MAE %", "Capture", "Diagnosis",
        "Risk %", "Hold min", "Source", "Paper only",
    ])
    for r, row in enumerate(rows, 2):
        vals = [
            row["created_at"], row["symbol"], row["okx_inst_id"], row["timeframe"],
            row["family"], row["side"], row["exit_mode"], row["status"],
            row["result"], row["net_pct"], row["net_r"], row["mfe_pct"],
            row["mae_pct"], row["capture"], row["diagnosis"], row["risk_pct"],
            row["max_hold_minutes"], row["source"], row["paper_only"],
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in (10, 11) and isinstance(val, (int, float)):
                if val > 0:
                    cell.fill = F_TP
                elif val < 0:
                    cell.fill = F_SL

    sr = len(rows) + 3
    summary = [
        ("Rows", len(rows)),
        ("Positive net", f'=COUNTIF(J2:J{max(sr - 2, 2)},">0")'),
        ("Negative net", f'=COUNTIF(J2:J{max(sr - 2, 2)},"<0")'),
        ("Avg net %", f'=IF(COUNTA(J2:J{max(sr - 2, 2)})=0,"",ROUND(AVERAGE(J2:J{max(sr - 2, 2)}),3))'),
        ("Avg net R", f'=IF(COUNTA(K2:K{max(sr - 2, 2)})=0,"",ROUND(AVERAGE(K2:K{max(sr - 2, 2)}),3))'),
    ]
    for i, (lbl, val) in enumerate(summary):
        ws.cell(row=sr + i, column=1, value=lbl).fill = F_SUM
        ws.cell(row=sr + i, column=1).font = FONT_BOLD
        ws.cell(row=sr + i, column=2, value=val).fill = F_SUM

    family_start = sr
    ws.cell(row=family_start, column=4, value="Family").fill = F_SUM
    ws.cell(row=family_start, column=5, value="Count").fill = F_SUM
    by_family: dict[str, int] = {}
    by_diag: dict[str, int] = {}
    by_result: dict[str, int] = {}
    for row in rows:
        by_family[row["family"]] = by_family.get(row["family"], 0) + 1
        by_diag[row["diagnosis"]] = by_diag.get(row["diagnosis"], 0) + 1
        by_result[row["result"]] = by_result.get(row["result"], 0) + 1
    for i, (name, count) in enumerate(sorted(by_family.items(), key=lambda kv: (-kv[1], kv[0]))[:20], 1):
        ws.cell(row=family_start + i, column=4, value=name)
        ws.cell(row=family_start + i, column=5, value=count)

    diag_start = sr
    ws.cell(row=diag_start, column=7, value="Diagnosis").fill = F_SUM
    ws.cell(row=diag_start, column=8, value="Count").fill = F_SUM
    for i, (name, count) in enumerate(sorted(by_diag.items(), key=lambda kv: (-kv[1], kv[0]))[:20], 1):
        ws.cell(row=diag_start + i, column=7, value=name)
        ws.cell(row=diag_start + i, column=8, value=count)

    result_start = sr
    ws.cell(row=result_start, column=10, value="Result").fill = F_SUM
    ws.cell(row=result_start, column=11, value="Count").fill = F_SUM
    for i, (name, count) in enumerate(sorted(by_result.items(), key=lambda kv: (-kv[1], kv[0]))[:20], 1):
        ws.cell(row=result_start + i, column=10, value=name)
        ws.cell(row=result_start + i, column=11, value=count)

    _set_widths(ws, [20, 12, 18, 7, 22, 8, 14, 14, 12, 9, 8, 8, 8, 8, 20, 8, 9, 12, 10])
    ws.freeze_panes = "A2"


# ── Entry point ───────────────────────────────────────────────────────────────

def build() -> None:
    screener  = _load_screener()
    main_ws   = _load_main_ws()
    bb_fade   = _load_bb_fade()
    simulator = _load_all_simulator_signals()
    pump      = _load_pump()
    impulse   = _load_impulse()
    main_impulse = _load_main_impulse()
    real      = _load_real()
    paper_watch = _load_paper_watch()

    wb = openpyxl.Workbook()
    _build_screener(wb, screener)
    _build_simulator(wb, simulator)
    _build_main_ws(wb, main_ws)
    _build_bb_fade(wb, bb_fade)
    _build_pump(wb, pump)
    _build_impulse(wb, impulse)
    _build_impulse(wb, main_impulse, "Main Impulse")
    _build_real(wb, real)
    _build_dashboard(wb)
    _build_charts(wb, main_ws, bb_fade, pump)
    _build_paper_watch(wb, paper_watch)

    wb.save(JOURNAL_PATH)
    n_labeled = sum(1 for r in screener if r["outcome"])
    print(
        f"Журнал готов: {JOURNAL_PATH}\n"
        f"  Скринер {len(screener)} ({n_labeled} закрыто) | "
        f"Main WS {len(main_ws)} | BB Fade {len(bb_fade)} | Main Impulse {len(main_impulse)} | "
        f"Памп {len(pump)} | Импульс {len(impulse)} | "
        f"Симулятор {len(simulator)} | Ручные {len(real)}"
    )


if __name__ == "__main__":
    build()
